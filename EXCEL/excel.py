import openpyxl  # Importing the openpyxl library to work with Excel files

# Load the Excel workbook named 'inventory.xlsx' which contains product details
inv_file = openpyxl.load_workbook("inventory.xlsx")

# Select the first sheet from the workbook which is assumed to contain the inventory data
product_list = inv_file["Sheet1"]

# Initializing dictionaries to aggregate and analyze the data
products_per_supplier = {}  # This dictionary will store how many products each supplier provides
total_value_per_supplier = {}  # This dictionary will store the total value of products each supplier provides
products_under_10_inv = {}  # This dictionary will identify products with low inventory (less than 10 items)

# Loop through each row in the sheet starting from the second row to skip the header
for product_row in range(2, product_list.max_row + 1):
    # Extracting data from the current row
    supplier_name = product_list.cell(product_row, 4).value  # Fetch the supplier's name
    inventory = product_list.cell(product_row, 2).value  # Fetch the inventory count for the product
    price = product_list.cell(product_row, 3).value  # Fetch the price of the product
    product_num = product_list.cell(product_row, 1).value  # Fetch the unique product number
    inventory_price = product_list.cell(product_row, 5)  # store the total value of the product's inventory

    # Update the count of products provided by each supplier
    if supplier_name in products_per_supplier:
        current_num_products = products_per_supplier.get(supplier_name)
        products_per_supplier[supplier_name] = current_num_products + 1
    else:
        products_per_supplier[supplier_name] = 1

    # Update the total value of products provided by each supplier
    if supplier_name in total_value_per_supplier:
        current_total_value = total_value_per_supplier.get(supplier_name)
        total_value_per_supplier[supplier_name] = current_total_value + inventory * price
    else:
        total_value_per_supplier[supplier_name] = inventory * price

    # Check if the product's inventory is below 10 and if so, add it to the products_under_10_inv dictionary
    if inventory < 10:
        products_under_10_inv[int(product_num)] = int(inventory)

    # Calculate the total inventory value for this product and store it back in the Excel sheet
    inventory_price.value = inventory * price

# Display the aggregated data
print("Number of products per supplier:", products_per_supplier)
print("Total inventory value per supplier:", total_value_per_supplier)
print("Products with less than 10 items in inventory:", products_under_10_inv)

# Save the Excel sheet with the added total inventory values for each product
inv_file.save("inventory_with_total_value.xlsx")
